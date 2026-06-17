"""
流程：
1) 通过搜索接口分页拿到文章列表
2) 逐篇进入详情页提取正文
3) 导出 CSV 和 Excel
"""

import csv
from datetime import datetime
import json
import os
import socket
import time
from typing import Dict, List

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from requests.exceptions import RequestException


SEARCH_API = "https://section.blog.naver.com/ajax/SearchList.naver"
KEYWORD = "서대문형무소역사관"
COUNT_PER_PAGE = 7
START_PAGE = 1
TARGET_COUNT = 1000

OUTPUT_CSV = "naver_blog_results.csv"
OUTPUT_EXCEL = "naver_blog_results.xlsx"

REQUEST_TIMEOUT = 15
MAX_RETRIES = 3
#代理
PROXY_URL = "http://127.0.0.1:7897"
AUTO_DETECT_LOCAL_PROXY = True

VERIFY_SSL = True


def build_headers() -> Dict[str, str]:
    """构造请求头，降低被反爬拦截概率。"""
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://section.blog.naver.com/",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,ko;q=0.7",
    }


def safe_get(session: requests.Session, url: str) -> requests.Response:
    """带重试的 GET 请求封装。"""
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT, verify=VERIFY_SSL)
            resp.raise_for_status()
            return resp
        except RequestException as exc:
            last_exc = exc
            print(f"[WARN] 第 {attempt}/{MAX_RETRIES} 次请求失败: {exc}")
            time.sleep(0.8)
    raise RequestException(f"请求失败，已重试 {MAX_RETRIES} 次: {last_exc}")


def is_port_open(host: str, port: int, timeout: float = 0.6) -> bool:
    """探测端口是否可连接，用于判断本机代理是否在运行。"""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(timeout)
        return sock.connect_ex((host, port)) == 0


def resolve_proxy_url() -> str:
    """按优先级获取可用代理地址。"""
    if PROXY_URL.strip():
        return PROXY_URL.strip()

    for key in (
        "HTTPS_PROXY",
        "https_proxy",
        "HTTP_PROXY",
        "http_proxy",
        "ALL_PROXY",
        "all_proxy",
    ):
        value = os.environ.get(key, "").strip()
        if value:
            return value

    if AUTO_DETECT_LOCAL_PROXY:
        candidates = [
            ("http://127.0.0.1:7897", "127.0.0.1", 7897),
            ("http://127.0.0.1:7890", "127.0.0.1", 7890),
            ("http://127.0.0.1:10809", "127.0.0.1", 10809),
            ("http://127.0.0.1:1080", "127.0.0.1", 1080),
        ]
        for proxy, host, port in candidates:
            if is_port_open(host, port):
                return proxy

    return ""


def build_search_url(page: int) -> str:
    """生成搜索分页 URL。"""
    return (
        f"{SEARCH_API}?countPerPage={COUNT_PER_PAGE}&currentPage={page}&endDate="
        f"&keyword={KEYWORD}&orderBy=sim&startDate=&type=post"
    )


def fetch_search_list(session: requests.Session, page: int) -> List[Dict]:
    """请求搜索接口并返回 searchList。"""
    try:
        resp = safe_get(session, build_search_url(page))
        raw = resp.text.strip()
        # 某些情况下返回值前会带防护前缀，先去掉再解析 JSON。
        if raw.startswith(")]}',"):
            raw = raw[5:].strip()
        data = json.loads(raw)
        search_list = data.get("result", {}).get("searchList", [])
        return search_list if isinstance(search_list, list) else []
    except RequestException as exc:
        print(f"[ERROR] 搜索接口请求失败（第 {page} 页）: {exc}")
        return []
    except ValueError as exc:
        snippet = resp.text[:300].replace("\n", " ").replace("\r", " ")
        print(f"[ERROR] 搜索接口 JSON 解析失败（第 {page} 页）: {exc}")
        print(f"[DEBUG] 响应片段: {snippet}")
        return []


def format_publish_time(raw_value: object) -> str:
    """将 addDate 统一格式化为可读时间字符串。"""
    if raw_value is None:
        return ""
    text = str(raw_value).strip()
    if not text:
        return ""
    try:
        if text.isdigit():
            ts = int(text)
            if len(text) >= 13:
                dt = datetime.fromtimestamp(ts / 1000)
            elif len(text) == 10:
                dt = datetime.fromtimestamp(ts)
            else:
                return text
            return dt.strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, OSError, OverflowError):
        return text
    return text


def extract_basic_info(item: Dict) -> Dict[str, str]:
    """提取单条搜索结果中的基础字段。"""
    def to_text(value: object) -> str:
        return "" if value is None else str(value).strip()

    # 获取title 
    title = BeautifulSoup(to_text(item.get("title")), "lxml").get_text(" ", strip=True)
    return {
        "title": title,
        "link": to_text(item.get("postUrl")),
        "summary": to_text(item.get("briefContents")),
        "publish_time": format_publish_time(item.get("addDate")),
        "blogger_name": to_text(item.get("nickName")),
    }


def parse_blog_content(html: str) -> str:
    """从博客页面 HTML 中提取正文文本。"""
    soup = BeautifulSoup(html, "lxml")
    container = soup.select_one("div.se-main-container")
    if container:
        return container.get_text("\n", strip=True)
    container = soup.select_one("#postViewArea, #content-area, div#post-area")
    if container:
        return container.get_text("\n", strip=True)
    return ""


def fetch_blog_content(session: requests.Session, post_url: str) -> str:
    """抓取博客详情正文，兼容 iframe 主体页面。"""
    if not post_url:
        return ""
    try:
        resp = safe_get(session, post_url)
    except RequestException as exc:
        print(f"[WARN] 详情页请求失败: {post_url} | {exc}")
        return ""
    soup = BeautifulSoup(resp.text, "lxml")

    iframe = soup.select_one("iframe#mainFrame")
    if iframe and iframe.get("src"):
        iframe_src = iframe["src"]
        iframe_url = f"https://blog.naver.com{iframe_src}" if iframe_src.startswith("/") else iframe_src
        try:
            iframe_resp = safe_get(session, iframe_url)
            return parse_blog_content(iframe_resp.text)
        except RequestException as exc:
            print(f"[WARN] iframe 请求失败: {iframe_url} | {exc}")
            return ""
    return parse_blog_content(resp.text)


def save_to_csv(rows: List[Dict[str, str]], file_name: str) -> None:
    """保存为 CSV，便于数据通用处理。"""
    fieldnames = ["title", "link", "summary", "publish_time", "blogger_name", "content"]
    with open(file_name, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def save_to_excel(rows: List[Dict[str, str]], file_name: str) -> None:
    """保存为更易读的 Excel（表头样式、冻结首行、自动列宽）。"""
    headers = [
        ("title", "文章标题"),
        ("link", "链接"),
        ("summary", "摘要"),
        ("publish_time", "发布时间"),
        ("blogger_name", "博主名称"),
        ("content", "正文内容"),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Naver博客结果"

    for col_idx, (_, header_name) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (field, _) in enumerate(headers, start=1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if field in {"summary", "content"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                cell.alignment = Alignment(vertical="center", horizontal="left")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        width = min(max(12, max_len + 2), 60)
        if col_letter == "F":
            width = 60
        ws.column_dimensions[col_letter].width = width

    wb.save(file_name)


def main() -> None:
    """主流程：配置代理 -> 分页抓取 -> 抓详情 -> 导出文件。"""
    session = requests.Session()
    session.headers.update(build_headers())

    proxy = resolve_proxy_url()
    if not proxy:
        print("[ERROR] 没检测到代理配置。请先设置 PROXY_URL。")
        return

    # 强制使用当前脚本解析出来的代理，避免系统环境变量干扰。
    session.trust_env = False
    session.proxies.update({"http": proxy, "https": proxy})
    print(f"[INFO] 当前使用代理: {proxy}")

    results: List[Dict[str, str]] = []
    seen_links = set()
    page = START_PAGE

    while len(results) < TARGET_COUNT:
        search_list = fetch_search_list(session, page)
        if not search_list:
            print(f"[INFO] 第 {page} 页没有结果了，结束。")
            break

        print(f"[INFO] 第 {page} 页：{len(search_list)} 条")
        page_new_count = 0

        for item in search_list:
            if len(results) >= TARGET_COUNT:
                break

            base = extract_basic_info(item)
            link = base["link"]
            # 按链接去重，避免重复记录。
            if not link or link in seen_links:
                continue

            seen_links.add(link)
            base["content"] = fetch_blog_content(session, link)
            results.append(base)
            page_new_count += 1

            print(f"[INFO] {len(results)}/{TARGET_COUNT}  {base['title']}")
            time.sleep(0.8)

        if page_new_count == 0:
            print("[INFO] 这一页没有新增文章，继续翻页也没意义，停。")
            break

        page += 1

    save_to_csv(results, OUTPUT_CSV)
    save_to_excel(results, OUTPUT_EXCEL)
    print(f"[INFO] 导出完成: {OUTPUT_CSV} / {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()

